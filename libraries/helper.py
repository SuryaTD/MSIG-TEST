import random
import string
from robot.api.deco import library, keyword
from datetime import datetime, timedelta
from openpyxl import load_workbook
import re
import xlrd
import csv
import json
from deepdiff import DeepDiff
from robot.libraries.BuiltIn import BuiltIn
from robot.api import logger

@library
class Utils:

    @keyword
    def generate_random_name(self, length=6):
        """Generate a random lowercase name of given length."""
        return ''.join(random.choices(string.ascii_lowercase, k=length))

    @keyword
    def generate_random_number(self, start=1000, end=9999):
        """Generate a random integer between start and end (inclusive)."""
        return random.randint(start, end)

    @keyword
    def update_submission_id(self, new_id, path="./testdata/data.py"):
        in_test1_block = False
        updated_lines = []

        with open(path, "r") as f:
            lines = f.readlines()

        for line in lines:
            stripped = line.strip()
            if stripped.startswith("NewUser"):
                in_test1_block = True
                updated_lines.append(line)
                continue

            if in_test1_block:
                if "submission_id" in stripped:
                    indent = line[:len(line) - len(line.lstrip())]
                    updated_lines.append(f'{indent}"submission_id": "{new_id}",\n')
                elif stripped.startswith("}"):
                    in_test1_block = False
                    updated_lines.append(line)
                else:
                    updated_lines.append(line)
            else:
                updated_lines.append(line)

        with open(path, "w") as f:
            f.writelines(updated_lines)

    @keyword
    def update_task_name(self, new_id, path="./testdata/data.py"):
        in_test1_block = False
        updated_lines = []

        with open(path, "r") as f:
            lines = f.readlines()

        for line in lines:
            stripped = line.strip()
            if stripped.startswith("TC_UI_281"):
                in_test1_block = True
                updated_lines.append(line)
                continue

            if in_test1_block:
                if "NewTaskName" in stripped:
                    indent = line[:len(line) - len(line.lstrip())]
                    updated_lines.append(f'{indent}"NewTaskName": "{new_id}",\n')
                elif stripped.startswith("}"):
                    in_test1_block = False
                    updated_lines.append(line)
                else:
                    updated_lines.append(line)
            else:
                updated_lines.append(line)

        with open(path, "w") as f:
            f.writelines(updated_lines)

    @keyword
    def update_task_id(self, new_id, path="./testdata/data.py"):
        in_test1_block = False
        updated_lines = []

        with open(path, "r") as f:
            lines = f.readlines()

        for line in lines:
            stripped = line.strip()
            if stripped.startswith("TC_UI_281"):
                in_test1_block = True
                updated_lines.append(line)
                continue

            if in_test1_block:
                if "data_task_id" in stripped:
                    indent = line[:len(line) - len(line.lstrip())]
                    updated_lines.append(f'{indent}"data_task_id": "{new_id}",\n')
                elif stripped.startswith("}"):
                    in_test1_block = False
                    updated_lines.append(line)
                else:
                    updated_lines.append(line)
            else:
                updated_lines.append(line)

        with open(path, "w") as f:
            f.writelines(updated_lines)

    @keyword
    def update_data(self, variable_name: str, key: str, new_value: str, path="./testdata/data.py"):
        in_variable_block = False
        updated_lines = []

        with open(path, "r") as f:
            lines = f.readlines()

        for line in lines:
            stripped = line.strip()

            # Detect the start of the dictionary (e.g., TC_UI_281 = { )
            if stripped.startswith(f"{variable_name}"):
                in_variable_block = True
                updated_lines.append(line)
                continue

            if in_variable_block:
                if f'"{key}"' in stripped and ':' in stripped:
                    indent = line[:len(line) - len(line.lstrip())]
                    updated_line = f'{indent}"{key}": "{new_value}",\n'
                    updated_lines.append(updated_line)
                elif stripped.startswith("}"):
                    in_variable_block = False
                    updated_lines.append(line)
                else:
                    updated_lines.append(line)
            else:
                updated_lines.append(line)

        with open(path, "w") as f:
            f.writelines(updated_lines)

    @keyword
    def get_tomorrows_date(self):
        tomorrow = datetime.today() + timedelta(days=1)
        return tomorrow.strftime('%m-%d-%Y')
    
    @keyword
    def get_tomorrows_date_YMD(self):
        tomorrow = datetime.today() + timedelta(days=1)
        return tomorrow.strftime('%Y-%m-%d')

    @keyword
    def get_tomorrows_date_format(self):
        tomorrow = datetime.today() + timedelta(days=1)
        return tomorrow.strftime('%m/%d/%Y')

    @keyword
    def get_formatted_current_date(self):
        return datetime.now().strftime("%a, %b %d %Y")

    @keyword
    def get_formatted_tomorrow_date(self):
        """Returns tomorrow's date in the format 'Day, Month DD YYYY' (e.g. 'Mon, Mar 25 2024')"""
        tomorrow = datetime.now() + timedelta(days=1)
        return tomorrow.strftime("%a, %b %d %Y")

    def normalize_value(self, val, ignore_case=False):
        """Normalize values by:
        - Stripping whitespace
        - Removing commas
        - Lowercasing (optional)
        - Converting float integers like 1.0 to '1'"""
        if val is None:
            return ''
        val_str = str(val).strip()
        val_str = re.sub(r',', '', val_str)
        try:
            # Convert '1.0' → '1'
            float_val = float(val_str)
            if float_val.is_integer():
                val_str = str(int(float_val))
            else:
                val_str = str(float_val)
        except ValueError:
            pass  # Not a number, keep as string
        if ignore_case:
            val_str = val_str.lower()
        return val_str

    @keyword
    def compare_excel_with_ui_list(self, excel_path, ui_data, sheet_name=None, ignore_case=False):
        """Compare Excel values with UI list. Fails test if mismatch is found."""
        excel_data = self._extract_excel_data(excel_path, sheet_name, ignore_case)
        ui_data_cleaned = [self.normalize_value(val, ignore_case) for val in ui_data if val not in (None, '', ' ')]

        excel_set = set(excel_data)
        ui_set = set(ui_data_cleaned)

        matched = [val for val in ui_data_cleaned if val in excel_set]
        missing = [val for val in excel_set if val not in ui_set]
        unexpected = [val for val in ui_set if val not in excel_set]

        # Log results in UI order
        log_lines = []
        if not missing and not unexpected:
            log_lines.append("✅ All values matched between Excel and UI.")
        else:
            log_lines.append("❌ Mismatch found between Excel and UI.")

        log_lines.append(f"Matched values ({len(matched)}): {matched}")
        if missing:
            log_lines.append(f"Missing in UI ({len(missing)}): {missing}")
        if unexpected:
            log_lines.append(f"Unexpected in UI ({len(unexpected)}): {unexpected}")

        logger.info("\n".join(log_lines))

        # Assert on mismatch
        if missing or unexpected:
            BuiltIn().fail("Excel and UI data mismatch found.")

    def _extract_excel_data(self, excel_path, sheet_name, ignore_case):
        """Extracts non-empty, normalized values from Excel sheet."""
        file_ext = excel_path.split('.')[-1].lower()
        data = []

        if file_ext == 'xlsx':
            with load_workbook(excel_path, data_only=True) as wb:
                ws = wb[sheet_name] if sheet_name else wb.active
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell not in (None, '', ' '):
                            data.append(self.normalize_value(cell, ignore_case))
        elif file_ext == 'xls':
            wb = xlrd.open_workbook(excel_path)
            ws = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
            for row_idx in range(ws.nrows):
                for cell in ws.row(row_idx):
                    if cell.value not in (None, '', ' '):
                        data.append(self.normalize_value(cell.value, ignore_case))
        else:
            raise ValueError("Unsupported file type. Only .xlsx and .xls are supported.")

        return data

    @keyword
    def compare_excel_files(self, actual_path, expected_path, sheet_name=None, ignore_case=False, ignore_columns=None):
        """
        Compare two Excel files content with column-wise comparison.
        Args:
            actual_path: Path to the actual Excel file
            expected_path: Path to the expected Excel file
            sheet_name: Name of the sheet to compare (optional)
            ignore_case: Whether to ignore case during comparison (default: False)
            ignore_columns: List of column names to ignore during comparison (case-insensitive)
        """
        # Extract data from both files with column information
        actual_data, actual_columns = self._extract_excel_data_with_columns(actual_path, sheet_name, ignore_case)
        expected_data, expected_columns = self._extract_excel_data_with_columns(expected_path, sheet_name, ignore_case)

        # Convert ignore_columns to lowercase for case-insensitive comparison
        ignore_columns_lower = []
        if ignore_columns:
            if isinstance(ignore_columns, str):
                ignore_columns_lower = [ignore_columns.lower()]
            else:
                ignore_columns_lower = [col.lower() for col in ignore_columns]

        # First check: Compare row counts
        actual_row_count = len(actual_data)
        expected_row_count = len(expected_data)
        
        if actual_row_count != expected_row_count:
            error_msg = f"\n❌ Row count mismatch:\nActual file rows: {actual_row_count}\nExpected file rows: {expected_row_count}"
            logger.error(error_msg)
            BuiltIn().fail(error_msg)

        # Second check: Compare column counts and names (excluding ignored columns)
        filtered_actual_columns = [col for col in actual_columns if col.lower() not in ignore_columns_lower]
        filtered_expected_columns = [col for col in expected_columns if col.lower() not in ignore_columns_lower]

        if len(filtered_actual_columns) != len(filtered_expected_columns):
            error_msg = f"\n❌ Column count mismatch (after ignoring specified columns):\nActual columns: {filtered_actual_columns}\nExpected columns: {filtered_expected_columns}"
            logger.error(error_msg)
            BuiltIn().fail(error_msg)

        # Create column-wise comparison results
        comparison_results = {}
        for col_idx, col_name in enumerate(actual_columns):
            # Skip ignored columns
            if col_name.lower() in ignore_columns_lower:
                continue

            actual_col_data = [row[col_idx] for row in actual_data if row]
            expected_col_data = [row[col_idx] for row in expected_data if row]
            
            matched = []
            missing = []
            unexpected = []
            
            for value in actual_col_data:
                if value in expected_col_data:
                    matched.append(value)
                else:
                    unexpected.append(value)
            
            for value in expected_col_data:
                if value not in actual_col_data:
                    missing.append(value)
            
            comparison_results[col_name] = {
                'matched': matched,
                'missing': missing,
                'unexpected': unexpected
            }

        # Log results in table format
        has_mismatch = False
        table_rows = []
        table_rows.append("| Column Name | Status | Details |")
        table_rows.append("|------------|--------|---------|")

        if ignore_columns_lower:
            table_rows.append(f"| NOTE | ℹ️ Info | Ignoring columns: {', '.join(ignore_columns_lower)} |")

        for col_name, result in comparison_results.items():
            if result['missing'] or result['unexpected']:
                has_mismatch = True
                status = "❌ MISMATCH"
                details = []
                if result['missing']:
                    details.append(f"Missing: {result['missing']}")
                if result['unexpected']:
                    details.append(f"Unexpected: {result['unexpected']}")
                details_str = "<br>".join(details)
            else:
                status = "✅ MATCH"
                details_str = f"All {len(result['matched'])} values match:<br>Values: {result['matched']}"
            
            table_rows.append(f"| {col_name} | {status} | {details_str} |")

        # Log the table
        if has_mismatch:
            logger.error("\n❌ Excel Comparison Results:\n" + "\n".join(table_rows))
            BuiltIn().fail("Excel files data mismatch found. Check the log for details.")
        else:
            logger.info("\n✅ Excel Comparison Results:\n" + "\n".join(table_rows))

    def _extract_excel_data_with_columns(self, excel_path, sheet_name, ignore_case):
        """
        Extracts data and column names from Excel sheet or CSV file.
        Returns tuple of (data_rows, column_names).
        """
        file_ext = excel_path.split('.')[-1].lower()
        data_rows = []
        columns = []

        if file_ext == 'csv':
            with open(excel_path, 'r', encoding='utf-8-sig') as csvfile:
                csvreader = csv.reader(csvfile)
                # Get column names from first row
                columns = next(csvreader)
                columns = [str(cell) if cell is not None else '' for cell in columns]
                
                # Get data rows (excluding header)
                for row in csvreader:
                    row_data = []
                    for cell in row:
                        if cell not in (None, '', ' '):
                            row_data.append(self.normalize_value(cell, ignore_case))
                        else:
                            row_data.append('')
                    if any(row_data):  # Only add rows that have at least one non-empty value
                        data_rows.append(row_data)

        elif file_ext == 'xlsx':
            wb = load_workbook(excel_path, data_only=True)
            try:
                ws = wb[sheet_name] if sheet_name else wb.active
                # Get column names from first row
                first_row = next(ws.iter_rows(values_only=True))
                columns = [str(cell) if cell is not None else '' for cell in first_row]
                
                # Get data rows (excluding header)
                for row in list(ws.iter_rows(min_row=2, values_only=True)):
                    row_data = []
                    for cell in row:
                        if cell not in (None, '', ' '):
                            row_data.append(self.normalize_value(cell, ignore_case))
                        else:
                            row_data.append('')
                    if any(row_data):  # Only add rows that have at least one non-empty value
                        data_rows.append(row_data)
            finally:
                wb.close()
        elif file_ext == 'xls':
            wb = xlrd.open_workbook(excel_path)
            ws = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
            
            # Get column names from first row
            columns = [str(cell.value) if cell.value is not None else '' for cell in ws.row(0)]
            
            # Get data rows (excluding header)
            for row_idx in range(1, ws.nrows):
                row_data = []
                for cell in ws.row(row_idx):
                    if cell.value not in (None, '', ' '):
                        row_data.append(self.normalize_value(cell.value, ignore_case))
                    else:
                        row_data.append('')
                if any(row_data):  # Only add rows that have at least one non-empty value
                    data_rows.append(row_data)
        else:
            raise ValueError("Unsupported file type. Only .xlsx, .xls, and .csv files are supported.")

        return data_rows, columns

    @keyword
    def compare_json_files(self, actual_path, expected_path, ignore_order=True, ignore_case=False):
        """
        Compare two JSON files and provide detailed comparison results.
        Handles large datasets (3000+ records) efficiently.
        
        Args:
            actual_path: Path to the actual JSON file
            expected_path: Path to the expected JSON file
            ignore_order: Whether to ignore list order during comparison (default: True)
            ignore_case: Whether to ignore string case during comparison (default: False)
        """
        try:
            with open(actual_path, 'r', encoding='utf-8') as f:
                actual_data = json.load(f)
            with open(expected_path, 'r', encoding='utf-8') as f:
                expected_data = json.load(f)
        except json.JSONDecodeError as e:
            error_msg = f"\n❌ Invalid JSON format: {str(e)}"
            logger.error(error_msg)
            BuiltIn().fail(error_msg)
        except FileNotFoundError as e:
            error_msg = f"\n❌ File not found: {str(e)}"
            logger.error(error_msg)
            BuiltIn().fail(error_msg)

        # First check: Compare total record counts
        actual_count = len(actual_data) if isinstance(actual_data, list) else 1
        expected_count = len(expected_data) if isinstance(expected_data, list) else 1

        logger.info(f"\n📊 Record Count Comparison:")
        logger.info(f"Actual file records: {actual_count}")
        logger.info(f"Expected file records: {expected_count}")

        if actual_count != expected_count:
            error_msg = f"\n❌ Record count mismatch: Actual ({actual_count}) != Expected ({expected_count})"
            logger.error(error_msg)
            BuiltIn().fail(error_msg)

        # Perform detailed comparison using DeepDiff
        diff_options = {
            'ignore_order': ignore_order,
            'ignore_string_case': ignore_case,
            'report_repetition': True,
            'verbose_level': 2
        }

        diff = DeepDiff(expected_data, actual_data, **diff_options)

        if diff:
            # Format and log differences in table format
            table_rows = []
            table_rows.append("\n| Field/Index | Status | Expected Value | Actual Value |")
            table_rows.append("|-------------|--------|----------------|--------------|")

            for change_type, changes in diff.items():
                if change_type == 'values_changed':
                    for path, change in changes.items():
                        field = self._get_field_name(path)
                        table_rows.append(f"| {field} | ❌ Changed | {change['old_value']} | {change['new_value']} |")
                
                elif change_type == 'dictionary_item_added':
                    for path in changes:
                        field = self._get_field_name(path)
                        value = self._get_value_at_path(actual_data, path)
                        table_rows.append(f"| {field} | ➕ Added | - | {value} |")
                
                elif change_type == 'dictionary_item_removed':
                    for path in changes:
                        field = self._get_field_name(path)
                        value = self._get_value_at_path(expected_data, path)
                        table_rows.append(f"| {field} | ❌ Removed | {value} | - |")
                
                elif change_type == 'iterable_item_added':
                    for path in changes:
                        field = self._get_field_name(path)
                        table_rows.append(f"| {field} | ➕ Added | - | {changes[path]} |")
                
                elif change_type == 'iterable_item_removed':
                    for path in changes:
                        field = self._get_field_name(path)
                        table_rows.append(f"| {field} | ❌ Removed | {changes[path]} | - |")
                
                elif change_type == 'type_changes':
                    for path, change in changes.items():
                        field = self._get_field_name(path)
                        table_rows.append(f"| {field} | ⚠️ Type Changed | {change['old_type']}: {change['old_value']} | {change['new_type']}: {change['new_value']} |")

            error_msg = "\n❌ JSON Comparison Results:\n" + "\n".join(table_rows)
            logger.error(error_msg)
            BuiltIn().fail("JSON files data mismatch found. Check the log for details.")
        else:
            logger.info("\n✅ JSON files match perfectly!")
            # Log matching data in table format for verification
            table_rows = []
            table_rows.append("\n| Field/Index | Status | Value |")
            table_rows.append("|-------------|--------|-------|")
            
            if isinstance(actual_data, list):
                for idx, item in enumerate(actual_data):
                    if isinstance(item, dict):
                        for key, value in item.items():
                            table_rows.append(f"| [{idx}].{key} | ✅ Match | {value} |")
                    else:
                        table_rows.append(f"| [{idx}] | ✅ Match | {item} |")
            elif isinstance(actual_data, dict):
                for key, value in actual_data.items():
                    table_rows.append(f"| {key} | ✅ Match | {value} |")
            else:
                table_rows.append(f"| root | ✅ Match | {actual_data} |")
            
            logger.info("\n".join(table_rows))

    def _get_field_name(self, path):
        """Extract readable field name from DeepDiff path."""
        # Remove 'root' and clean up the path
        clean_path = path.replace('root', '').strip('[]').replace('\'', '')
        parts = clean_path.split('][')
        
        # Format the field name
        if not parts[0]:
            return "root"
        
        result = []
        for part in parts:
            if part.isdigit():
                result.append(f"[{part}]")
            else:
                if result:
                    result.append(f".{part}")
                else:
                    result.append(part)
        
        return ''.join(result)

    def _get_value_at_path(self, data, path):
        """Get value at a specific path in the JSON structure"""
        try:
            # Remove 'root' from the path and split into parts
            parts = path.replace('root', '').strip('[]').split('][')
            current = data
            
            for part in parts:
                # Remove quotes from string keys
                part = part.strip('\'\"')
                if part.isdigit():
                    current = current[int(part)]
                else:
                    current = current[part]
            
            return current
        except (KeyError, IndexError, TypeError):
            return "N/A"

    @keyword
    def get_json_value_by_query(self, json_file_path, query_path):
        """
        Load a JSON file and get value using a dot-notation query path.
        Automatically detects and handles both single values and lists/arrays.
        
        Example:
        | ${value}= | Get Json Value By Query | path/to/file.json | d3Submission.d3Company.naics.0.code |
        | ${value}= | Get Json Value By Query | data.json | user.address.city |
        | @{list}= | Get Json Value By Query | data.json | items.*.name |  # Gets all item names
        
        Query Path Examples and Return Types:
        | Query Path                              | Return Type | Description                    |
        | d3Submission.d3Company.name             | Single     | Direct property access         |
        | d3Submission.d3Company.naics.0.code     | Single     | Array index access            |
        | d3Submission.d3Company.naics.*.code     | List       | All codes from naics array    |
        | d3Submission.items.*.details.*.name     | List       | Nested array traversal        |
        
        Args:
            json_file_path: Path to the JSON file
            query_path: Query path in dot notation. Can be with or without quotes.
                       Use '*' for array/list traversal
                       (e.g., "d3Submission.d3Company.naics.*.code" for all codes)
            
        Returns:
            - Single value if the path points to a specific element
            - List of values if the path contains wildcards or points to an array
            - None if path not found or error occurs
        """
        # Remove single quotes from the query path if present
        query_path = query_path.replace("'", "")
        try:
            # Load the JSON file
            with open(json_file_path, 'r') as f:
                data = json.load(f)
            
            # Log the structure type we're working with
            logger.info(f"Loaded JSON data type: {type(data)}")
            if isinstance(data, dict):
                logger.info(f"Top level keys: {list(data.keys())}")
            elif isinstance(data, list):
                logger.info(f"Array length: {len(data)}")
            
            # Split the query path and handle array indices and wildcards
            path_elements = []
            contains_wildcard = False
            for element in query_path.split('.'):
                # Check if it's a number (array index)
                if element == '*':
                    contains_wildcard = True
                    path_elements.append(element)
                elif element.isdigit():
                    path_elements.append(int(element))
                else:
                    path_elements.append(element)
            
            def traverse_path(current_data, remaining_path):
                if not remaining_path:
                    return [current_data]
                
                element = remaining_path[0]
                rest_path = remaining_path[1:]
                
                if element == '*':
                    if isinstance(current_data, list):
                        results = []
                        for item in current_data:
                            results.extend(traverse_path(item, rest_path))
                        return results
                    else:
                        logger.error(f"Wildcard '*' used on non-list element. Current data type: {type(current_data)}")
                        return None
                else:
                    try:
                        next_data = current_data[element]
                        # Log the type of data we're traversing
                        logger.debug(f"Accessing element '{element}' -> type: {type(next_data)}")
                        return traverse_path(next_data, rest_path)
                    except (KeyError, IndexError) as e:
                        logger.error(f"Path element '{element}' not found: {str(e)}")
                        logger.error(f"Available keys: {list(current_data.keys()) if isinstance(current_data, dict) else 'N/A'}")
                        return None
            
            results = traverse_path(data, path_elements)
            
            if not results:
                logger.warn(f"No results found for path '{query_path}'")
                return None
            elif len(results) == 1 and not contains_wildcard:
                logger.info(f"Single value result for '{query_path}' -> Type: {type(results[0])}, Value: {results[0]}")
                return results[0]
            else:
                logger.info(f"Multiple values result for '{query_path}' -> Count: {len(results)}, Types: {[type(r) for r in results]}")
                logger.info(f"Values: {results}")
                return results
            
        except FileNotFoundError:
            logger.error(f"JSON file not found: {json_file_path}")
            return None
        except json.JSONDecodeError:
            logger.error(f"Invalid JSON file: {json_file_path}")
            return None
        except (KeyError, IndexError) as e:
            logger.error(f"Path '{query_path}' not found in JSON: {str(e)}")
            return None
        except Exception as e:
            logger.error(f"Error processing JSON query: {str(e)}")
            return None

    @keyword
    def get_current_day_number(self, time_zone='local', increment='0'):
        """Returns only the current day number (e.g., 28 for July 28).
 
        Arguments:
        - time_zone: 'local' (default) or 'UTC'.
        - increment: Optional time increment in format 'HH:MM:SS', can be negative.
        """
        # Get current datetime based on timezone
        if time_zone.upper() == 'LOCAL':
            dt = datetime.now()
        elif time_zone.upper() == 'UTC':
            try:
                dt = datetime.now(datetime.UTC).replace(tzinfo=None)  # Python 3.12+
            except AttributeError:
                dt = datetime.utcnow()
        else:
            raise ValueError(f"Unsupported timezone '{time_zone}'")
 
        # Apply increment if provided
        if increment != '0':
            h, m, s = [int(part) for part in increment.split(':')]
            dt += timedelta(hours=h, minutes=m, seconds=s)
 
        return dt.day